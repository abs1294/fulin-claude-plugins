# MSPDI XML → XLSX（MS Project 匯出格式對等 + 階層可讀性排版）
#
# 目的：免去「XML → 匯入 MS Project → 存 MPP → 另存 XLSX」這一圈手工，
#       直接從 MSPDI XML 產出「內容與 MS Project 匯出完全一致」的 XLSX，
#       並額外加上 MS Project 匯出所沒有的可讀性排版（階層縮排、層級配色、凍結窗格）。
#
# 內容對等契約（逐欄比對 MS Project 匯出結果驗證過，改動前先想清楚）：
#   A 大綱編號   1 / 1.1 / 1.1.1（依 OutlineLevel 累進，非 UID）
#   B 名稱       任務名原文（排版縮排用 alignment.indent，不動字串內容）
#   C 工期       "N 工作日"（Duration PT{h}H → h/8，整除去小數）
#   D 開始時間   "2026年7月27日 上午 08:00"（月/日不補零；12 小時制 + 上午/下午）
#   E 完成時間   同上
#   F 前置任務   PredecessorLink → 對應列的大綱編號（無則空白）
#   G 資源名稱   Assignment 對應的 Resource Name（摘要列空白）
#   H 完成百分比 PercentComplete，預設 0
#
# ⚠ 內容欄位不得為了「好看」而改寫（例如把工期寫成 "3 天"、日期改 ISO），
#   否則與客戶端 MS Project 匯出結果對不上。排版只用樣式層（字體/底色/縮排/欄寬）。
import re, sys, html
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def iso_to_date(ts):
    """2026-07-27T08:00:00 → datetime.date(2026,7,27)；無值回 None。"""
    if not ts:
        return None
    m = re.match(r'(\d{4})-(\d{2})-(\d{2})', ts)
    return date(int(m.group(1)), int(m.group(2)), int(m.group(3))) if m else None

NS = '{http://schemas.microsoft.com/project}'


def parse_tasks(xml_text):
    """取出 Task 節點的必要欄位（用正則而非 XML parser，避免 namespace 雜訊）。"""
    tasks = []
    for tm in re.finditer(r'<Task>(.*?)</Task>', xml_text, re.S):
        b = tm.group(1)
        def g(tag):
            m = re.search(r'<%s>(.*?)</%s>' % (tag, tag), b, re.S)
            return html.unescape(m.group(1)) if m else None
        preds = re.findall(r'<PredecessorUID>(\d+)</PredecessorUID>', b)
        tasks.append({
            'uid': g('UID'), 'name': g('Name') or '',
            'level': int(g('OutlineLevel') or 1),
            'start': g('Start'), 'finish': g('Finish'),
            'dur': g('Duration'), 'summary': g('Summary') == '1',
            'milestone': g('Milestone') == '1',
            'pct': g('PercentComplete') or '0',
            'preds': preds,
        })
    return tasks


def parse_resources(xml_text):
    res = {}
    for rm in re.finditer(r'<Resource>(.*?)</Resource>', xml_text, re.S):
        b = rm.group(1)
        uid = re.search(r'<UID>(\d+)</UID>', b)
        name = re.search(r'<Name>(.*?)</Name>', b, re.S)
        if uid and name:
            res[uid.group(1)] = html.unescape(name.group(1))
    assign = {}
    for am in re.finditer(r'<Assignment>(.*?)</Assignment>', xml_text, re.S):
        b = am.group(1)
        t = re.search(r'<TaskUID>(\d+)</TaskUID>', b)
        r = re.search(r'<ResourceUID>(\d+)</ResourceUID>', b)
        if t and r and r.group(1) in res:
            assign.setdefault(t.group(1), []).append(res[r.group(1)])
    return assign


def fmt_duration(dur):
    """PT80H0M0S → '10 工作日'（MS Project 以 8 小時為一工作日）。"""
    if not dur:
        return ''
    m = re.match(r'PT(\d+)H', dur)
    if not m:
        return ''
    days = int(m.group(1)) / 8
    s = str(int(days)) if days == int(days) else str(days)
    return '%s 工作日' % s


def fmt_datetime(ts):
    """2026-07-27T08:00:00 → '2026年7月27日'（只到日；本專案不細控管時間）。"""
    if not ts:
        return ''
    m = re.match(r'(\d{4})-(\d{2})-(\d{2})', ts)
    if not m:
        return ''
    y, mo, d = m.groups()
    return '%s年%s月%s日' % (y, int(mo), int(d))


def build_outline_numbers(tasks):
    """依 OutlineLevel 累進產生 1 / 1.1 / 1.1.1。"""
    counters, out = [], []
    for t in tasks:
        lv = t['level']
        if lv > len(counters):
            counters.extend([0] * (lv - len(counters)))
        counters = counters[:lv]
        counters[lv - 1] += 1
        out.append('.'.join(str(c) for c in counters))
    return out


def convert(xml_path, xlsx_path):
    with open(xml_path, encoding='utf-8-sig') as f:
        xml_text = f.read()

    tasks = parse_tasks(xml_text)
    assign = parse_resources(xml_text)
    outline = build_outline_numbers(tasks)
    uid2outline = {t['uid']: outline[i] for i, t in enumerate(tasks)}

    wb = Workbook()
    ws = wb.active
    ws.title = '專案時程'

    # 註：本專案不細控管時間，日期欄只到日，故標題用「開始日期／完成日期」
    #     （MS Project 原生匯出為「開始時間／完成時間」，此處刻意改稱以符實際內容）
    # D/E「開始/完成日期」存**真 Excel 日期**，用儲存格格式顯示成「2026年10月23日」——
    #   如此畫面是中文日期、底層仍是可運算日期，I 欄公式可直接引用 D/E（免額外輔助欄）。
    # I 欄「預計完成%」：依開檔當天(TODAY())算「今天照計畫應完成到幾趴」的基準線，
    #   對應使用者 MPP 自訂欄位公式（工作天粒度：已消耗工作天 ÷ 總工作天，clip 0–100%）。
    DATE_FMT = 'yyyy"年"m"月"d"日"'   # 真日期，中文顯示
    # 欄序（使用者指定）：預計完成%(H) 在前、完成百分比(I) 在後，兩欄皆以 % 顯示。
    #   H 預計完成% = 依開檔日算的計畫基準線（活公式）
    #   I 完成百分比 = 實際完成度（XML PercentComplete，目前 0）
    headers = ['大綱編號', '名稱', '工期', '開始日期', '完成日期', '前置任務', '資源名稱',
               '預計完成%', '完成百分比']
    ws.append(headers)

    # ── 排版樣式（MS Project 匯出所沒有的可讀性強化）──────────────
    HDR_FILL = PatternFill('solid', fgColor='1F3864')      # 深藍表頭
    L1_FILL = PatternFill('solid', fgColor='D9E2F3')       # L1 模組：淺藍
    L2_FILL = PatternFill('solid', fgColor='F2F2F2')       # L2 原文條目：淺灰
    MS_FILL = PatternFill('solid', fgColor='FFF2CC')       # 里程碑：淺黃
    thin = Side(style='thin', color='BFBFBF')
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in range(1, 10):   # A–I 表頭套樣式；J/K 隱藏輔助欄不套
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name='Microsoft JhengHei', size=11, bold=True, color='FFFFFF')
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
    ws.row_dimensions[1].height = 24

    for i, t in enumerate(tasks):
        r = i + 2
        preds = '、'.join(uid2outline.get(p, '') for p in t['preds'] if uid2outline.get(p))
        res_name = '' if t['summary'] else '、'.join(assign.get(t['uid'], []))
        ws.cell(row=r, column=1, value=outline[i])
        ws.cell(row=r, column=2, value=t['name'])
        ws.cell(row=r, column=3, value=fmt_duration(t['dur']))
        sd = iso_to_date(t['start'])
        fd = iso_to_date(t['finish'])
        cS = ws.cell(row=r, column=4, value=sd)   # D 開始日期（真日期＋中文格式）
        cF = ws.cell(row=r, column=5, value=fd)   # E 完成日期
        if sd:
            cS.number_format = DATE_FMT
        if fd:
            cF.number_format = DATE_FMT
        ws.cell(row=r, column=6, value=preds)
        ws.cell(row=r, column=7, value=res_name)

        # H 欄：預計完成%（活公式，開檔當天重算）。摘要/里程碑列不算（留空）。
        # 直接引用 D/E（真日期）：TODAY()<=起→0；TODAY()>=迄→1；否則 已消耗工作天/總工作天。
        #   NETWORKDAYS 含頭尾，故 -1 得「已過完的工作天」；分母同樣 -1；除零時給 1。
        if not t['summary'] and sd and fd:
            f = ('=IFERROR(IF(TODAY()<=D{r},0,IF(TODAY()>=E{r},1,'
                 'MAX(0,NETWORKDAYS(D{r},TODAY())-1)/MAX(1,NETWORKDAYS(D{r},E{r})-1))),0)').format(r=r)
            ch = ws.cell(row=r, column=8, value=f)
            ch.number_format = '0%'

        # I 欄：完成百分比（實際完成度，XML PercentComplete）→ 以 % 顯示（存小數，格式化）
        ci = ws.cell(row=r, column=9, value=int(t['pct']) / 100)
        ci.number_format = '0%'

        lv = t['level']
        bold = t['summary'] or t['milestone']
        fill = None
        if t['milestone']:
            fill = MS_FILL
        elif lv == 1:
            fill = L1_FILL
        elif t['summary']:
            fill = L2_FILL

        for c in range(1, 10):   # A–I 套樣式；J/K 為隱藏輔助欄不套
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name='Microsoft JhengHei', size=10, bold=bold,
                             color='1F3864' if lv == 1 else '000000')
            if fill:
                cell.fill = fill
            cell.border = BORDER
            if c == 2:
                # 階層以縮排呈現（不改動名稱字串本身）
                cell.alignment = Alignment(indent=(lv - 1) * 2, vertical='center', wrap_text=False)
            elif c in (1, 3, 4, 5, 6, 8, 9):
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')

        # MS Project 式群組（可摺疊）：L2 縮 1 階、L3 縮 2 階
        if lv > 1:
            ws.row_dimensions[r].outlineLevel = lv - 1

    widths = {'A': 11, 'B': 68, 'C': 11, 'D': 26, 'E': 26, 'F': 11, 'G': 22, 'H': 12, 'I': 11}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'C2'                      # 凍結表頭 + 大綱編號/名稱兩欄
    ws.auto_filter.ref = 'A1:I%d' % (len(tasks) + 1)
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_view.showGridLines = False
    wb.save(xlsx_path)
    return len(tasks)


if __name__ == '__main__':
    src = sys.argv[1]
    dst = sys.argv[2]
    n = convert(src, dst)
    print('OK 轉出 %d 列 → %s' % (n, dst))
